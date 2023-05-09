from PIL import Image
from openpyxl import load_workbook

# Open xlsx file and store the file_name column in a list
wb = load_workbook('humanoid_images.xlsx')
ws = wb.active
file_name = []
for i in range(2, ws.max_row + 1):
    file_name.append(ws.cell(row=i, column=4).value)


# split the image name and ignore .jpg from its name
for i in range(len(file_name)):
    file_name[i] = file_name[i].split('.')[0]

# execute the code for each file in the list
# for i in file_name:
for i in range(len(file_name)):
    temp = file_name[i].split('.')[0]
    # Open the image
    img = Image.open(temp + '.jpg')

    # Get the original size of the image
    original_size = img.size

    # Set the desired size of the cropped image
    desired_size = (550, 550)

    # Calculate the amount to crop from each edge
    left = (original_size[0] - desired_size[0]) // 2
    top = (original_size[1] - desired_size[1]) // 2
    right = original_size[0] - left - desired_size[0]
    bottom = original_size[1] - top - desired_size[1]

    # left = 760
    # top = 340
    # right = 760
    # bottom = 340

    # Crop the image
    img = img.crop((left, top, original_size[0] - right, original_size[1] - bottom))

    # Save the cropped image
    img.save(temp + '.jpg')
